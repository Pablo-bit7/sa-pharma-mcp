"""
Tests for mpr_utils.py – link discovery, caching, DataFrame parsing, and
fallback behaviour for the Database of Medicine Prices (MPR/SEP).
"""
import io
import pytest
import pandas as pd
import httpx
from unittest.mock import AsyncMock, MagicMock, patch


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_response(content=b"", text="", status_code=200):
    resp = MagicMock(spec=httpx.Response)
    resp.status_code = status_code
    resp.content = content
    resp.text = text
    resp.raise_for_status = MagicMock()
    return resp


def _minimal_mpr_excel() -> bytes:
    """
    Build an in-memory Excel file that mimics the MPR structure.
    Data starts at row 2 (header=1 for pandas).
    Target columns: 1,6,7,10,11,12,3,13,14,16 (0-indexed).
    We create 17 columns so all target indices are valid.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1 is skipped metadata; row 2 is the header.
    ws.append(["skip_row"])

    num_cols = 17
    headers = [f"col_{i}" for i in range(num_cols)]
    ws.append(headers)

    # One data row – recognisable values at the target column positions.
    # Positions: 1=Applicant, 6=Prop_Name, 7=Active_Ingr, 10=Dosage_Form,
    #            11=Pack_Size, 12=Quantity, 3=NAPPI_Code, 13=Mfr_Price,
    #            14=Logistics_Fee, 16=SEP
    row = [""] * num_cols
    row[1] = "Aspen"            # Applicant
    row[3] = "NAP001"           # NAPPI_Code
    row[6] = "Panado"           # Proprietery_Name
    row[7] = "Paracetamol"      # Active_Ingredients (non-null → row kept)
    row[10] = "Tablet"          # Dosage_Form
    row[11] = "100"             # Pack_Size
    row[12] = "500 mg"          # Quantity
    row[13] = "2.50"            # Manufacturer_Price
    row[14] = "0.10"            # Logistics_Fee
    row[16] = "2.65"            # SEP
    ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


NHI_HTML_WITH_LINK = (
    '<a href="/wp-content/uploads/database-of-prices-2024.xlsx">'
    "Medicine Prices</a>"
)

NHI_HTML_WITHOUT_LINK = "<html><body>Nothing here</body></html>"


# ---------------------------------------------------------------------------
# Tests – discover_latest_mpr_list_link
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_discover_mpr_link_returns_absolute_url():
    import mpr_utils

    mock_client = AsyncMock(spec=httpx.AsyncClient)
    mock_client.get.return_value = _make_response(text=NHI_HTML_WITH_LINK)

    url = await mpr_utils.discover_latest_mpr_list_link(mock_client)

    assert url.startswith("http")
    assert "database-of-prices-2024.xlsx" in url


@pytest.mark.asyncio
async def test_discover_mpr_link_raises_when_not_found():
    import mpr_utils

    mock_client = AsyncMock(spec=httpx.AsyncClient)
    mock_client.get.return_value = _make_response(text=NHI_HTML_WITHOUT_LINK)

    with pytest.raises(ValueError, match="Database of Medicine Prices"):
        await mpr_utils.discover_latest_mpr_list_link(mock_client)


@pytest.mark.asyncio
async def test_discover_mpr_link_case_insensitive_prices():
    """Pattern should match URLs containing 'prices' in any case."""
    import mpr_utils

    html = '<a href="/files/Prices_2024.xlsx">Download</a>'
    mock_client = AsyncMock(spec=httpx.AsyncClient)
    mock_client.get.return_value = _make_response(text=html)

    url = await mpr_utils.discover_latest_mpr_list_link(mock_client)
    assert "Prices_2024.xlsx" in url


# ---------------------------------------------------------------------------
# Tests – get_latest_mpr_list_df
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_get_mpr_df_uses_cache_when_link_unchanged(tmp_path):
    """If the stored link matches the current link, return the cached CSV."""
    import mpr_utils

    link = "https://example.com/database-of-prices.xlsx"

    cache_csv = tmp_path / "ndoh_mpr_sep_cache.csv"
    cache_df = pd.DataFrame({"Applicant": ["Aspen"], "Active_Ingredients": ["Paracetamol"]})
    cache_df.to_csv(cache_csv, index=False)

    link_file = tmp_path / "ndoh_mpr_latest_link.txt"
    link_file.write_text(link)

    with (
        patch.object(mpr_utils, "CACHE_FILE", str(cache_csv)),
        patch.object(mpr_utils, "LINK_TRACKER", str(link_file)),
        patch("mpr_utils.discover_latest_mpr_list_link", new=AsyncMock(return_value=link)),
    ):
        df = await mpr_utils.get_latest_mpr_list_df()

    assert not df.empty
    assert "Applicant" in df.columns


@pytest.mark.asyncio
async def test_get_mpr_df_downloads_when_link_changed(tmp_path):
    """When a new link is found, the Excel file is downloaded and cache rebuilt."""
    import mpr_utils

    new_link = "https://example.com/database-of-prices-NEW.xlsx"
    old_link = "https://example.com/database-of-prices-OLD.xlsx"

    link_file = tmp_path / "ndoh_mpr_latest_link.txt"
    link_file.write_text(old_link)
    cache_csv = tmp_path / "ndoh_mpr_sep_cache.csv"

    excel_bytes = _minimal_mpr_excel()
    excel_response = _make_response(content=excel_bytes)

    with (
        patch.object(mpr_utils, "CACHE_FILE", str(cache_csv)),
        patch.object(mpr_utils, "LINK_TRACKER", str(link_file)),
        patch("mpr_utils.discover_latest_mpr_list_link", new=AsyncMock(return_value=new_link)),
        patch("httpx.AsyncClient") as mock_cls,
    ):
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get.return_value = excel_response
        mock_cls.return_value = mock_client

        df = await mpr_utils.get_latest_mpr_list_df()

    assert not df.empty
    assert "Active_Ingredients" in df.columns
    assert cache_csv.exists()
    assert link_file.read_text() == new_link


@pytest.mark.asyncio
async def test_get_mpr_df_falls_back_to_stale_cache_on_error(tmp_path):
    """When the network fails and a cache file exists, it must be returned."""
    import mpr_utils

    cache_csv = tmp_path / "ndoh_mpr_sep_cache.csv"
    stale_df = pd.DataFrame({"Applicant": ["StaleApplicant"], "Active_Ingredients": ["StaleIngr"]})
    stale_df.to_csv(cache_csv, index=False)

    link_file = tmp_path / "ndoh_mpr_latest_link.txt"

    with (
        patch.object(mpr_utils, "CACHE_FILE", str(cache_csv)),
        patch.object(mpr_utils, "LINK_TRACKER", str(link_file)),
        patch("mpr_utils.discover_latest_mpr_list_link", new=AsyncMock(side_effect=RuntimeError("network down"))),
    ):
        df = await mpr_utils.get_latest_mpr_list_df()

    assert df["Applicant"].iloc[0] == "StaleApplicant"


@pytest.mark.asyncio
async def test_get_mpr_df_raises_when_no_cache_and_network_fails(tmp_path):
    """When no cache exists and the network fails, the exception propagates."""
    import mpr_utils

    cache_csv = tmp_path / "ndoh_mpr_sep_cache.csv"  # does NOT exist
    link_file = tmp_path / "ndoh_mpr_latest_link.txt"

    with (
        patch.object(mpr_utils, "CACHE_FILE", str(cache_csv)),
        patch.object(mpr_utils, "LINK_TRACKER", str(link_file)),
        patch("mpr_utils.discover_latest_mpr_list_link", new=AsyncMock(side_effect=RuntimeError("network down"))),
    ):
        with pytest.raises(RuntimeError, match="network down"):
            await mpr_utils.get_latest_mpr_list_df()


@pytest.mark.asyncio
async def test_get_mpr_df_drops_rows_with_null_active_ingredients(tmp_path):
    """Rows where Active_Ingredients is NaN must be dropped during parsing."""
    import mpr_utils
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["skip"])
    ws.append([f"col_{i}" for i in range(17)])

    row_good = [""] * 17
    row_good[7] = "Paracetamol"
    row_good[1] = "Aspen"
    ws.append(row_good)

    row_bad = [""] * 17
    row_bad[7] = None   # no Active_Ingredients → will be dropped
    row_bad[1] = "Cipla"
    ws.append(row_bad)

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    new_link = "https://example.com/database-of-prices.xlsx"
    link_file = tmp_path / "ndoh_mpr_latest_link.txt"
    link_file.write_text("https://old.link")
    cache_csv = tmp_path / "ndoh_mpr_sep_cache.csv"

    with (
        patch.object(mpr_utils, "CACHE_FILE", str(cache_csv)),
        patch.object(mpr_utils, "LINK_TRACKER", str(link_file)),
        patch("mpr_utils.discover_latest_mpr_list_link", new=AsyncMock(return_value=new_link)),
        patch("httpx.AsyncClient") as mock_cls,
    ):
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get.return_value = _make_response(content=excel_bytes)
        mock_cls.return_value = mock_client

        df = await mpr_utils.get_latest_mpr_list_df()

    assert len(df) == 1
    assert df["Active_Ingredients"].iloc[0] == "Paracetamol"


@pytest.mark.asyncio
async def test_get_mpr_df_forward_fills_non_ingredient_columns(tmp_path):
    """
    The MPR has merged cells for applicant/product; ffill() propagates values
    down to rows that only carry a new active ingredient.
    """
    import mpr_utils
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["skip"])
    ws.append([f"col_{i}" for i in range(17)])

    # First data row – has all values.
    row1 = [""] * 17
    row1[1] = "Aspen"
    row1[6] = "Panado"
    row1[7] = "Paracetamol"
    row1[13] = "2.50"
    ws.append(row1)

    # Second row – Applicant/Proprietery_Name is blank (simulates merged cell);
    # only Active_Ingredients differs.
    row2 = [""] * 17
    row2[1] = None      # blank → must be ffilled from row1
    row2[6] = None
    row2[7] = "Codeine"
    row2[13] = "3.00"
    ws.append(row2)

    buf = io.BytesIO()
    wb.save(buf)

    new_link = "https://example.com/database-of-prices.xlsx"
    link_file = tmp_path / "ndoh_mpr_latest_link.txt"
    link_file.write_text("https://old.link")
    cache_csv = tmp_path / "ndoh_mpr_sep_cache.csv"

    with (
        patch.object(mpr_utils, "CACHE_FILE", str(cache_csv)),
        patch.object(mpr_utils, "LINK_TRACKER", str(link_file)),
        patch("mpr_utils.discover_latest_mpr_list_link", new=AsyncMock(return_value=new_link)),
        patch("httpx.AsyncClient") as mock_cls,
    ):
        mock_client = AsyncMock()
        mock_client.__aenter__ = AsyncMock(return_value=mock_client)
        mock_client.__aexit__ = AsyncMock(return_value=False)
        mock_client.get.return_value = _make_response(content=buf.getvalue())
        mock_cls.return_value = mock_client

        df = await mpr_utils.get_latest_mpr_list_df()

    assert len(df) == 2
    # The forward-filled value from the first row should appear on the second.
    assert df["Applicant"].iloc[1] == "Aspen"
    assert df["Proprietery_Name"].iloc[1] == "Panado"
